import matplotlib.pyplot as plt
import numpy as np

import matplotlib as mpl

from os import listdir, makedirs 
import math
import openpyxl
from openpyxl.chart import (
    LineChart,
    Reference,
)
import openpyxl.chart
from time import sleep

path = r"C:/Users\AlexF\Desktop"
experiment = 'эксперемент 1'
frequency = '3.5'
gain = '273'
number_read_data = 0

def show_plt(bit0, bit1):
    fig, ax = plt.subplots()
    ax.plot(bit0)
    ax.plot(bit1)
    plt.show()

def finde_file(name_file):
    files = listdir("C:/Users/AlexF/Desktop/")
    for file in files:
        if file == name_file:
            return False
    return True

def read_data(filename="sin0"):
    result = list()
    while True:
        try:
            with open(f"{path}/{filename}.txt", "r+") as file:
                print(file.readline()[:-1])
                result = [int(line.rstrip('\n')) for line in file ]
        except(FileNotFoundError):
            continue
        else:
            break
    open(f"{path}/{filename}.txt", "w").close() # Очищаем файл
    print(result)
    with open(f'C:/Users/AlexF/Desktop/испытания/эксперимент{experiment}/{frequency}/{gain}/rawData_{filename}.txt', 'w', encoding="utf-8") as file:
        for res in result:
            file.write(str(res))
            file.write('\n')
    return result
    

def math_data(sin_raw, cos_raw):
    sin = list()
    cos = list()

    [sin.append(sin_raw[index]-2048) for index in range(0,64-1)]
    [cos.append(cos_raw[index]-2048) for index in range(0,64-1)]

    return [math.sqrt(sin[index]**2 + cos[index]**2) for index in range(0,64-1)]

def write_data_exl(bit0, bit1):
    xlsx = openpyxl.Workbook()
    sheet = xlsx.active
    sheet['A1'] = 'bit0'
    sheet['B1'] = 'bit1'
    for row in range(len(bit0)):
        sheet.cell(row=row+2, column=1).value = bit0[row]
        sheet.cell(row=row+2, column=2).value = bit1[row]
    
    values = Reference(sheet, min_col = 1, min_row = 2, max_col = 2, max_row = 64-1)
    chart = LineChart() 
    chart.add_data(values)
    chart.title = " LINE-CHART "
    chart.x_axis.title = " X-AXIS "
    chart.y_axis.title = " Y-AXIS "
    sheet.add_chart(chart, "E2")
    xlsx.save(f'C:/Users/AlexF/Desktop/испытания/эксперимент{experiment}/{frequency}/{gain}/result.xlsx')

if __name__ == '__main__':
    while True:
        experiment = input('Введите номер эксперемента: ')
        frequency = input('Введите частоту эксперемента: ')
        gain = input('Введите коэффициент усиления: ')
        makedirs(f'C:/Users/AlexF/Desktop/испытания/эксперимент{experiment}/{frequency}/{gain}') 
        while finde_file('sin0.txt'):
            pass
        while finde_file("cos0.txt"):
            pass
        while finde_file("sin1.txt"):
            pass
        while finde_file("cos1.txt"):
            pass

        sleep(2)

        sin0 = read_data("sin0")
        cos0 = read_data("cos0")
        sin1 = read_data("sin1")
        cos1 = read_data("cos1")


        bit0 = math_data(sin0, cos0)
        bit1 = math_data(sin1, cos1)
        write_data_exl(bit0, bit1)
        show_plt(bit0, bit1)